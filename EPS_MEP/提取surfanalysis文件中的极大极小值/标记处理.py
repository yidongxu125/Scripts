# version 1.01
# by yidongxu
# 用法：当使用16脚本提取极大值和极小值如果有的极小值小于5个就会出现提取错误的现象，改脚本运行后会对存在这个现象的文件进行标注，您在根据17脚本的excel文件进行单独的修改。
import os
import openpyxl
from openpyxl.styles import PatternFill

# 打开当前目录下所有的 Excel 文件
for filename in os.listdir():
    if filename.endswith('.xlsx'):
        wb = openpyxl.load_workbook(filename)
        # 循环处理 Excel 文件中的每个工作表
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            # 获取每行数据中第二列和第三列的值，如果小于5则填充为黄色
            for row in sheet.iter_rows(min_row=2, min_col=2, max_col=3):
                col2_val = int(row[0].value) if row[0].value else None
                col3_val = int(row[1].value) if row[1].value else None
                if col2_val is not None and col3_val is not None:
                    if col2_val < 5 or col3_val < 5:
                        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                        for cell in row:
                            cell.fill = yellow_fill
        # 保存修改后的 Excel 文件
        wb.save(filename)
